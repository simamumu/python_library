import pandas as pd
import os
import openpyxl
from collections import defaultdict,deque

# 調節用パラメータ
filename = 'kaito7.xlsx' # 入力ファイル名
Output = 'output7.xlsx' # 出力ディレクトリ


wb = openpyxl.load_workbook(filename)
sheets = wb.sheetnames

days = []
names = []
dict = defaultdict(dict)
for sheet in sheets:
    sh = wb[sheet]
    i = 3
    while True:
        tmp = sh.cell(row=1,column=i).value
        if tmp:
            days.append(tmp)
        else:
            break
        i += 1
    print(days)
    days.pop()

    i = 2
    while True:
        tmp = sh.cell(row=i,column=2).value
        if tmp:
            names.append(tmp)
        else:
            break
        i += 1

    W = len(days)
    H = len(names)
    for y in range(2,2+H):
        for x in range(3,3+W):
            tmp = sh.cell(row=y,column=x).value
            dict[names[y-2]][days[x-3]] = tmp

times = dict['しまむら']['7/10(水)'].split(', ')

ans = [[' ', ' '] + names]
for d in days:
    for t in times:
        tmpl = [d,t]
        for n in names:
            if dict[n][d] and t in dict[n][d]:
                tmpl.append(1)
            else:
                tmpl.append(0)
        ans.append(tmpl)

for a in ans:
    print(a)


wb = openpyxl.load_workbook(Output)
sheets = wb.sheetnames
sheet = wb[sheets[0]]


def write_list_2d(sheet, l_2d, start_row, start_col):
    for y, row in enumerate(l_2d):
        for x, cell in enumerate(row):
            #print(l_2d[y][x])
            sheet.cell(row=start_row + y,column=start_col + x,value=l_2d[y][x])
            #print(sheet.cell(row=start_row + y,column=start_col + x).value)

write_list_2d(sheet,ans,1,1)

wb.save(Output)

print(sheets[0])
